#!/usr/bin/env python3
"""
Build (or append to) a Customer Agreements Summary workbook from structured
extraction JSON, matching the final reviewed format for the Jacksonville
customer-agreements project (see ../references/conventions.md for the full
data model and writing rules this implements).

Usage:
    python3 build_workbook.py \
        --template "Jacksonville - Customer Agreements Summary.xlsx" \
        --data customers.json \
        --sheet-name "Full Draft" \
        --out output.xlsx

customers.json schema:
{
  "as_of_date": "YYYY-MM-DD",
  "banner": "optional custom banner text for row 1",
  "subtitle": "optional custom subtitle for row 2",
  "customers": [
    {
      "customer_id": 1131,
      "customer_name": "Speed Guide, Inc",
      "escalation_rate": 0.03,          // decimal fraction, or null/omit if unknown (defaults to 0%)
      "header_notes": "",                // only for customers with zero usable rows
      "rows": [
        // kind "msa": a Master Service Agreement / lease / governing-contract line.
        // No dates go in L/M (Beginning/Ending) or N (period MRC) - those columns
        // stay blank on msa rows. The FIRST msa row in a customer's row list is
        // where escalation_rate gets written (K column) and becomes the anchor
        // every so_recurring row's formula references.
        {"kind": "msa", "desc": "...", "agreement_date": "YYYY-MM-DD"|null,
         "renewal": "short phrase"|null, "holdover": "short phrase"|null,
         "payment": "short phrase"|null, "notes": "..."},

        // kind "so_fixed": a bounded, non-renewing commitment. Written as ONE
        // static row with an explicit begin+end. If the as-of date later moves
        // past `end` with no matching renewal row, this naturally drops out of
        // Current Total - that's intentional, not a bug to fix.
        {"kind": "so_fixed", "desc": "...", "agreement_date": "YYYY-MM-DD"|null,
         "renewal": "short phrase"|null, "mrc_commencement": "YYYY-MM-DD"|null,
         "mrc_expiration": "YYYY-MM-DD"|null, "begin": "YYYY-MM-DD",
         "end": "YYYY-MM-DD", "total_mrc": 1755.00,
         "categories": {"FC": 235, "PWRP": 960}, "notes": "..."},

        // kind "so_recurring": an auto-renewing commitment with no confirmed
        // end. Written as a FIRST static row (begin + total_mrc), then a chain
        // of formula-driven period rows (EDATE + compounding escalation off
        // the customer's escalation anchor) generated far enough forward to
        // cover as_of_date with room to spare. This is what makes "current
        // total" stay accurate as time passes without anyone re-running the
        // extraction.
        {"kind": "so_recurring", "desc": "...", "agreement_date": "YYYY-MM-DD"|null,
         "renewal": "short phrase"|null, "mrc_commencement": "YYYY-MM-DD"|null,
         "mrc_expiration": null, "begin": "YYYY-MM-DD", "total_mrc": 104.00,
         "categories": {"CC": 99, "PRTF": 5}, "notes": "..."},

        // kind "open": no defined start or end at all (e.g. a base/anchor line
        // item referenced by later orders). Always counted in Current Total
        // regardless of as_of_date.
        {"kind": "open", "desc": "...", "total_mrc": 167.72,
         "categories": {"CC": 50}, "notes": "..."}
      ],
      "current_total_notes": "explanation shown on the Current Total row"
    }
  ]
}

Category codes (column letter in the template):
  PWRFIX(O) CG(P) FC(Q) CC(R) PWRP(S) PWRR(T) PRTF(U) SMF(V) IPA(W) INET(X)
  BAS(Y) XCN(Z) PAN(AA) VPN(AB) RRS(AC) MST(AD) FBR(AE) BCS(AF) INTRA(AG)
"""
import argparse
import json
import re
from copy import copy
from datetime import datetime, timedelta

CODE_COL = {"PWRFIX": 15, "CG": 16, "FC": 17, "CC": 18, "PWRP": 19, "PWRR": 20,
            "PRTF": 21, "SMF": 22, "IPA": 23, "INET": 24, "BAS": 25, "XCN": 26,
            "PAN": 27, "VPN": 28, "RRS": 29, "MST": 30, "FBR": 31, "BCS": 32,
            "INTRA": 33}
NOTES_COL = 34  # AH

HEADER_SRC = 7
AGMT_TEXT_ONLY_SRC = 8
SO_FIRST_SRC = 9
BLANK_SRC = 13
CURTOTAL_SRC = 26


def norm_code(c):
    return re.sub(r"[^A-Z]", "", c.upper())


def parse_date(s):
    if not s:
        return None
    if isinstance(s, datetime):
        return s
    try:
        return datetime.strptime(s, "%Y-%m-%d")
    except Exception:
        return None


def map_cats(catdict):
    out, unmapped = {}, []
    for k, v in (catdict or {}).items():
        col = CODE_COL.get(norm_code(k))
        if col:
            out[col] = out.get(col, 0) + v
        else:
            unmapped.append(f"{k}={v}")
    return out, unmapped


def safe(s):
    """Never let a note/description start with a character openpyxl or Excel
    will interpret as a formula trigger (=, +, -, @). Prepend a space instead
    of stripping - keeps the text intact and human-readable."""
    if isinstance(s, str) and s[:1] in ("=", "+", "-", "@"):
        return " " + s
    return s


def build_workbook(template_path, data, sheet_title, out_path, append_to=None):
    import openpyxl
    from openpyxl.styles import Font, Alignment

    if append_to:
        wb = openpyxl.load_workbook(append_to)
    else:
        wb = openpyxl.load_workbook(template_path)

    src_tmpl_wb = openpyxl.load_workbook(template_path)
    src_ws = src_tmpl_wb["Customers (Scarlett)"]

    ws = wb.create_sheet(sheet_title)

    def copy_cell_style(dst, src):
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format

    for col_letter, dim in src_ws.column_dimensions.items():
        ws.column_dimensions[col_letter].width = dim.width
    ws.column_dimensions["AH"].width = 70

    banner = data.get("banner") or f"FULL DRAFT — {sheet_title}, extracted from source folders only"
    subtitle = data.get("subtitle") or (
        "Read the Notes column (AH) — every uncertain / unconfirmed / ambiguous field is flagged there. "
        "Change the MRC As-of Date below and Current Total recalculates automatically."
    )
    ws.cell(row=1, column=2, value=banner).font = Font(name="Times New Roman", size=12, bold=True, color="C00000")
    ws.cell(row=2, column=2, value=subtitle).font = Font(name="Calibri", size=10, italic=True)

    as_of = parse_date(data.get("as_of_date")) or datetime.today()
    ws.cell(row=3, column=2, value="MRC As-of Date:").font = Font(name="Calibri", size=11, bold=True)
    c3 = ws.cell(row=3, column=3, value=as_of)
    c3.number_format = "mm-dd-yy"

    HEADER_LABEL_ROW = 4
    for c in range(1, 34):
        s = src_ws.cell(row=6, column=c)
        d = ws.cell(row=HEADER_LABEL_ROW, column=c, value=s.value)
        copy_cell_style(d, s)
    d = ws.cell(row=HEADER_LABEL_ROW, column=NOTES_COL, value="Notes / Review Flags")
    copy_cell_style(d, src_ws.cell(row=6, column=33))
    ws.row_dimensions[HEADER_LABEL_ROW].height = src_ws.row_dimensions[6].height
    ws.freeze_panes = "B5"

    def style_row(dst_row, src_row, ncols=33):
        ws.row_dimensions[dst_row].height = src_ws.row_dimensions[src_row].height
        for c in range(2, ncols + 1):
            s = src_ws.cell(row=src_row, column=c)
            copy_cell_style(ws.cell(row=dst_row, column=c), s)
        sN = src_ws.cell(row=src_row, column=3)
        dN = ws.cell(row=dst_row, column=NOTES_COL)
        dN.font = copy(sN.font)
        dN.border = copy(sN.border)
        dN.alignment = Alignment(wrap_text=True, vertical="top")

    def write_blank(r):
        style_row(r, BLANK_SRC)

    row = 6
    write_blank(5)

    for cust in data.get("customers", []):
        cid = cust.get("customer_id")
        cname = cust.get("customer_name") or ""
        crows = cust.get("rows") or []
        header_notes = cust.get("header_notes") or ""
        esc_rate = cust.get("escalation_rate")

        if not crows:
            style_row(row, HEADER_SRC)
            ws.cell(row=row, column=2, value=cid)
            ws.cell(row=row, column=3, value=safe(cname))
            ws.cell(row=row, column=4, value=0)
            ws.cell(row=row, column=NOTES_COL, value=safe(header_notes or "No usable source documents found."))
            row += 2
            continue

        header_row = row
        style_row(row, HEADER_SRC)
        ws.cell(row=row, column=2, value=cid)
        ws.cell(row=row, column=3, value=safe(cname))
        if header_notes:
            ws.cell(row=row, column=NOTES_COL, value=safe(header_notes))
        row += 1
        block_first_data_row = row
        esc_anchor_row = None

        for rdict in crows:
            kind = rdict.get("kind")

            if kind == "msa":
                style_row(row, AGMT_TEXT_ONLY_SRC)
                ws.cell(row=row, column=3, value=safe(rdict.get("desc") or "Master Service Agreement"))
                ad = parse_date(rdict.get("agreement_date"))
                if ad:
                    ws.cell(row=row, column=5, value=ad)
                if rdict.get("renewal"):
                    ws.cell(row=row, column=6, value=safe(rdict["renewal"]))
                if rdict.get("holdover"):
                    ws.cell(row=row, column=7, value=safe(rdict["holdover"]))
                if rdict.get("payment"):
                    ws.cell(row=row, column=8, value=safe(rdict["payment"]))
                if rdict.get("notes"):
                    ws.cell(row=row, column=NOTES_COL, value=safe(rdict["notes"]))
                if esc_anchor_row is None:
                    esc_anchor_row = row
                    ws.cell(row=row, column=11, value=esc_rate if esc_rate is not None else None)
                    if esc_rate is not None:
                        ws.cell(row=row, column=11).number_format = "0%"
                row += 1

            elif kind == "so_fixed":
                cats, unmapped = map_cats(rdict.get("categories"))
                note = rdict.get("notes") or ""
                if unmapped:
                    note = (note + f" [Unmapped category codes kept in notes rather than dropped: {', '.join(unmapped)}]").strip()
                style_row(row, SO_FIRST_SRC)
                ws.cell(row=row, column=3, value=safe(rdict.get("desc") or "Sales Order"))
                ad = parse_date(rdict.get("agreement_date"))
                if ad:
                    ws.cell(row=row, column=5, value=ad)
                if rdict.get("renewal"):
                    ws.cell(row=row, column=6, value=safe(rdict["renewal"]))
                mc = parse_date(rdict.get("mrc_commencement"))
                me = parse_date(rdict.get("mrc_expiration"))
                if mc:
                    ws.cell(row=row, column=9, value=mc)
                if me:
                    ws.cell(row=row, column=10, value=me)
                begin = parse_date(rdict["begin"])
                end = parse_date(rdict["end"])
                ws.cell(row=row, column=12, value=begin)
                ws.cell(row=row, column=13, value=end)
                ws.cell(row=row, column=14, value=rdict.get("total_mrc") or 0)
                for col, val in cats.items():
                    ws.cell(row=row, column=col, value=val)
                if note:
                    ws.cell(row=row, column=NOTES_COL, value=safe(note))
                row += 1

            elif kind == "so_recurring":
                cats, unmapped = map_cats(rdict.get("categories"))
                note = rdict.get("notes") or ""
                if unmapped:
                    note = (note + f" [Unmapped category codes kept in notes rather than dropped: {', '.join(unmapped)}]").strip()
                first_row = row
                style_row(row, SO_FIRST_SRC)
                ws.cell(row=row, column=3, value=safe(rdict.get("desc") or "Sales Order"))
                ad = parse_date(rdict.get("agreement_date"))
                if ad:
                    ws.cell(row=row, column=5, value=ad)
                if rdict.get("renewal"):
                    ws.cell(row=row, column=6, value=safe(rdict["renewal"]))
                mc = parse_date(rdict.get("mrc_commencement"))
                if mc:
                    ws.cell(row=row, column=9, value=mc)
                begin = parse_date(rdict["begin"])
                ws.cell(row=row, column=12, value=begin)
                ws.cell(row=row, column=13, value=f"=EDATE(L{row},12)-1")
                ws.cell(row=row, column=14, value=rdict.get("total_mrc") or 0)
                for col, val in cats.items():
                    ws.cell(row=row, column=col, value=val)
                if note:
                    ws.cell(row=row, column=NOTES_COL, value=safe(note))
                row += 1

                if esc_anchor_row is None:
                    esc_anchor_row = first_row
                    if esc_rate is not None:
                        ws.cell(row=first_row, column=11, value=esc_rate)
                        ws.cell(row=first_row, column=11).number_format = "0%"

                cover_through = as_of + timedelta(days=366)
                cur_begin = begin
                periods_written = 0
                while cur_begin < cover_through and periods_written < 40:
                    style_row(row, SO_FIRST_SRC, ncols=14)
                    ws.cell(row=row, column=12, value=f"=M{row - 1}+1")
                    ws.cell(row=row, column=13, value=f"=EDATE(L{row},12)-1")
                    ws.cell(row=row, column=14, value=f"=N{row - 1}*(1+$K${esc_anchor_row})")
                    cur_begin = cur_begin.replace(year=cur_begin.year + 1)
                    row += 1
                    periods_written += 1

            elif kind == "open":
                cats, unmapped = map_cats(rdict.get("categories"))
                note = rdict.get("notes") or ""
                if unmapped:
                    note = (note + f" [Unmapped category codes kept in notes rather than dropped: {', '.join(unmapped)}]").strip()
                style_row(row, SO_FIRST_SRC)
                ws.cell(row=row, column=3, value=safe(rdict.get("desc") or "Ongoing item"))
                ws.cell(row=row, column=14, value=rdict.get("total_mrc") or 0)
                for col, val in cats.items():
                    ws.cell(row=row, column=col, value=val)
                if note:
                    ws.cell(row=row, column=NOTES_COL, value=safe(note))
                row += 1

        block_last_data_row = row - 1
        write_blank(row)
        row += 1
        ct_row = row
        style_row(row, CURTOTAL_SRC)
        ws.cell(row=row, column=13, value="Current Total")
        formula = (
            f'=IF(SUMPRODUCT((($L${block_first_data_row}:$L${block_last_data_row}="")'
            f'+($L${block_first_data_row}:$L${block_last_data_row}<=$C$3)>0)'
            f'*(($M${block_first_data_row}:$M${block_last_data_row}="")'
            f'+($M${block_first_data_row}:$M${block_last_data_row}>=$C$3)>0)'
            f'*ISNUMBER($N${block_first_data_row}:$N${block_last_data_row}))=0,"N/A",'
            f'SUMPRODUCT((($L${block_first_data_row}:$L${block_last_data_row}="")'
            f'+($L${block_first_data_row}:$L${block_last_data_row}<=$C$3)>0)'
            f'*(($M${block_first_data_row}:$M${block_last_data_row}="")'
            f'+($M${block_first_data_row}:$M${block_last_data_row}>=$C$3)>0)'
            f'*ISNUMBER($N${block_first_data_row}:$N${block_last_data_row})'
            f'*$N${block_first_data_row}:$N${block_last_data_row}))'
        )
        ws.cell(row=row, column=14, value=formula)
        cur_notes = cust.get("current_total_notes") or ""
        if cur_notes:
            ws.cell(row=row, column=NOTES_COL, value=safe(cur_notes))
        ws.cell(row=header_row, column=4, value=f"=N{ct_row}")
        row += 2

    for s in wb.worksheets:
        s.sheet_view.tabSelected = (s.title == sheet_title)
    wb.active = wb.sheetnames.index(sheet_title)
    ws.sheet_view.topLeftCell = "A1"
    ws.sheet_view.selection[0].activeCell = "B1"
    ws.sheet_view.selection[0].sqref = "B1"

    wb.save(out_path)
    return out_path, row


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--template", required=True, help="Path to the original master template xlsx")
    ap.add_argument("--data", required=True, help="Path to customers.json")
    ap.add_argument("--sheet-name", required=True, help="Name for the new sheet")
    ap.add_argument("--out", required=True, help="Output xlsx path")
    ap.add_argument("--append-to", default=None, help="Existing workbook to add the sheet to, instead of the template")
    args = ap.parse_args()

    data = json.load(open(args.data))
    out_path, last_row = build_workbook(args.template, data, args.sheet_name, args.out, append_to=args.append_to)
    print(f"Wrote {out_path} — sheet '{args.sheet_name}' through row {last_row}")


if __name__ == "__main__":
    main()
